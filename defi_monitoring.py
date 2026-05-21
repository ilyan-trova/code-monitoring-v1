# =============================================================================
#  MONITORING DeFi — Données on-chain directes  v7
#  Morpho Blue | Aave V3
#
#  LANCEMENT :
#  python3 defi_monitoring_v7.py        → continu, toutes les 5 min
#
# =============================================================================

import os, csv, sys, json, math, time
from datetime import datetime
from dotenv import load_dotenv
from web3 import Web3
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
#  CONFIGURATION
# =============================================================================

INTERVALLE_SECONDES   = 300
SEUIL_UTIL_ALERTE     = 90.0
SEUIL_APY_OPPORTUNITE = 6.0
TVL_MIN_M             = 1.0
FICHIER_EXCEL         = "monitoring_defi.xlsx"
FICHIER_CSV           = "historique_onchain.csv"
SECONDES_PAR_AN       = 365 * 24 * 3600

MORPHO_BLUE_ADDR   = "0xBBBBBbbBBb9cC5e90e3b3Af64bdAF62C37EEFFCb"
MORPHO_IRM_ADDR    = "0x870aC11D48B15DB9a138Cf899d20F13F79Ba00BC"
AAVE_PROVIDER_ADDR = "0x7B4EB56E7CD4b454BA8ff71E4518426369a138a3"
USDC_ADDR          = "0xA0b86991c6218b36c1d19D4a2e9Eb0cE3606eB48"


# =============================================================================
#  MARCHÉS MORPHO À SURVEILLER
#  Seul le market_id et le nom sont nécessaires.
#  Les paramètres (loanToken, collateralToken, oracle, lltv) sont lus
#  automatiquement on-chain via idToMarketParams().
# =============================================================================

MORPHO_MARCHES = {
    "WBTC/USDC (86%)"    : "b323495f7e4148be5643a4ea4a8221eef163e4bccfdedc2a6f4696baacbc86cc",
    "WETH/USDC (94.5%)"  : "3a85e619751152991742810df6ec69ce473daef99e28a64ab2340d7b7ccfee49",
    "cbBTC/USDC"         : "9103c3b4e834476c9a62ea009ba2c884ee42e94e6e314a26f04d312434191836",
    "wstETH/USDT"        : "e7e9694b754c4d4f7e21faf7223f6fa71abaeb10296a4c43a54a7977149687d2",
    "wstETH/USDC"        : "13c42741a359ac4a8aa8287d2be109dcf28344484f91185f9a79bd5a805a55ae",
    "cbETH/USDC"         : "1c21c59df9db44bf6f645d854ee710a8ca17b479451447e9f56758aee10a2fad",
}


# =============================================================================
#  ABIs
# =============================================================================

# Morpho Blue : lire l'état d'un marché
ABI_MORPHO_MARKET = json.loads('''[{
    "name": "market", "type": "function", "stateMutability": "view",
    "inputs": [{"name": "id", "type": "bytes32"}],
    "outputs": [{"type": "tuple", "components": [
        {"name": "totalSupplyAssets", "type": "uint128"},
        {"name": "totalSupplyShares", "type": "uint128"},
        {"name": "totalBorrowAssets", "type": "uint128"},
        {"name": "totalBorrowShares", "type": "uint128"},
        {"name": "lastUpdate",        "type": "uint128"},
        {"name": "fee",               "type": "uint128"}
    ]}]
}]''')

# Morpho Blue : lire les paramètres d'un marché depuis son ID
ABI_MORPHO_PARAMS = json.loads('''[{
    "name": "idToMarketParams", "type": "function", "stateMutability": "view",
    "inputs": [{"name": "id", "type": "bytes32"}],
    "outputs": [{"type": "tuple", "components": [
        {"name": "loanToken",       "type": "address"},
        {"name": "collateralToken", "type": "address"},
        {"name": "oracle",          "type": "address"},
        {"name": "irm",             "type": "address"},
        {"name": "lltv",            "type": "uint256"}
    ]}]
}]''')

# Morpho IRM : calculer le borrow rate
ABI_MORPHO_IRM = json.loads('''[{
    "name": "borrowRateView", "type": "function", "stateMutability": "view",
    "inputs": [
        {"name": "marketParams", "type": "tuple", "components": [
            {"name": "loanToken",       "type": "address"},
            {"name": "collateralToken", "type": "address"},
            {"name": "oracle",          "type": "address"},
            {"name": "irm",             "type": "address"},
            {"name": "lltv",            "type": "uint256"}
        ]},
        {"name": "market", "type": "tuple", "components": [
            {"name": "totalSupplyAssets", "type": "uint128"},
            {"name": "totalSupplyShares", "type": "uint128"},
            {"name": "totalBorrowAssets", "type": "uint128"},
            {"name": "totalBorrowShares", "type": "uint128"},
            {"name": "lastUpdate",        "type": "uint128"},
            {"name": "fee",               "type": "uint128"}
        ]}
    ],
    "outputs": [{"name": "borrowRate", "type": "uint256"}]
}]''')

# Aave V3 PoolDataProvider
ABI_AAVE = json.loads('''[{
    "name": "getReserveData", "type": "function", "stateMutability": "view",
    "inputs": [{"name": "asset", "type": "address"}],
    "outputs": [{"type": "tuple", "components": [
        {"name": "unbacked",                "type": "uint256"},
        {"name": "accruedToTreasuryScaled", "type": "uint256"},
        {"name": "totalAToken",             "type": "uint256"},
        {"name": "totalStableDebt",         "type": "uint256"},
        {"name": "totalVariableDebt",       "type": "uint256"},
        {"name": "liquidityRate",           "type": "uint256"},
        {"name": "variableBorrowRate",      "type": "uint256"},
        {"name": "stableBorrowRate",        "type": "uint256"},
        {"name": "averageStableBorrowRate", "type": "uint256"},
        {"name": "liquidityIndex",          "type": "uint256"},
        {"name": "variableBorrowIndex",     "type": "uint256"},
        {"name": "lastUpdateTimestamp",     "type": "uint40"}
    ]}]
}]''')

# ERC-20 : lire les décimales d'un token
ABI_ERC20 = json.loads('''[{
    "name": "decimals", "type": "function", "stateMutability": "view",
    "inputs": [], "outputs": [{"type": "uint8"}]
}]''')


# =============================================================================
#  CONVERSION TAUX
# =============================================================================

def wad_to_apy(rate_wad):
    """Morpho borrowRate WAD/seconde → APY %"""
    if rate_wad <= 0:
        return 0.0
    return (math.exp((rate_wad / 1e18) * SECONDES_PAR_AN) - 1) * 100

def ray_to_apy(rate_ray):
    """Aave taux annuel RAY (1e27) → APY %"""
    return (rate_ray / 1e27) * 100

def supply_apy(b_apy, util, fee=0):
    """APY prêteur = APY emprunteur × utilisation × (1 - fee)"""
    return b_apy * util * (1 - fee / 1e18)

def get_decimals(w3, token_addr):
    """Lit les décimales d'un token ERC-20 directement on-chain."""
    try:
        token = w3.eth.contract(
            address=Web3.to_checksum_address(token_addr),
            abi=ABI_ERC20
        )
        return token.functions.decimals().call()
    except:
        return 6  # fallback USDC par défaut


# =============================================================================
#  LECTURE MORPHO BLUE
# =============================================================================

def lire_morpho(w3, morpho, morpho_params, irm):
    """
    Pour chaque market_id configuré :
    1. Lit les paramètres du marché via idToMarketParams() (automatique)
    2. Lit l'état du marché via market()
    3. Lit les décimales du loan token via decimals()
    4. Calcule le borrow rate via borrowRateView()
    5. Dérive le supply APY et l'utilisation
    """
    resultats = []

    for nom, market_id_hex in MORPHO_MARCHES.items():
        try:
            market_id = bytes.fromhex(market_id_hex)

            # Appel 1 : paramètres du marché (loanToken, collateralToken, oracle, irm, lltv)
            params = morpho_params.functions.idToMarketParams(market_id).call()
            loan_token       = params[0]
            collateral_token = params[1]
            oracle           = params[2]
            irm_addr         = params[3]
            lltv             = params[4]

            # Appel 2 : état du marché (supply, borrow, fee)
            m            = morpho.functions.market(market_id).call()
            total_supply = m[0]
            total_borrow = m[2]
            fee_wad      = m[5]

            # Appel 3 : décimales du loan token (lu on-chain, pas hardcodé)
            dec = get_decimals(w3, loan_token)

            tvl_m = total_supply / 10**dec / 1e6

            # Ignore uniquement les marchés complètement vides (market_id invalide)
            if total_supply == 0:
                continue

            util = total_borrow / total_supply

            # Appel 4 : borrow rate via l'IRM
            # Les params sont passés exactement tels que retournés par idToMarketParams
            borrow_rate = irm.functions.borrowRateView(
                (loan_token, collateral_token, oracle, irm_addr, lltv),
                (m[0], m[1], m[2], m[3], m[4], m[5])
            ).call()

            b = wad_to_apy(borrow_rate)
            p = supply_apy(b, util, fee_wad)

            resultats.append({
                "protocole"   : "Morpho Blue",
                "marche"      : nom,
                "supply_apy"  : round(p, 2),
                "borrow_apy"  : round(b, 2),
                "utilisation" : round(util * 100, 1),
                "tvl_m"       : round(tvl_m, 2),
                "liquidite_m" : round((total_supply - total_borrow) / 10**dec / 1e6, 2),
                "tvl_faible"  : tvl_m < TVL_MIN_M,
            })

        except Exception as e:
            # Marché ignoré silencieusement si erreur (market_id invalide, pool vide...)
            pass

    return resultats


# =============================================================================
#  LECTURE AAVE V3
# =============================================================================

def lire_aave(w3, aave):
    resultats = []
    try:
        d      = aave.functions.getReserveData(
            Web3.to_checksum_address(USDC_ADDR)
        ).call()
        supply = d[2]
        borrow = d[4]
        util   = borrow / supply if supply > 0 else 0

        resultats.append({
            "protocole"   : "Aave V3",
            "marche"      : "USDC",
            "supply_apy"  : round(ray_to_apy(d[5]), 2),
            "borrow_apy"  : round(ray_to_apy(d[6]), 2),
            "utilisation" : round(util * 100, 1),
            "tvl_m"       : round(supply / 1e6 / 1e6, 2),
            "liquidite_m" : round((supply - borrow) / 1e6 / 1e6, 2),
            "tvl_faible"  : False,
        })
    except Exception as e:
        print(f"  ⚠️  Aave V3 : {e}")
    return resultats


# =============================================================================
#  AFFICHAGE TERMINAL
# =============================================================================

def afficher(resultats, bloc, timestamp):
    print(f'\n╔{"═"*82}╗')
    print(f'║  MONITORING DeFi — On-chain direct'
          f'{"":>18}Bloc {bloc:,}  [{timestamp}]  ║')
    print(f'╚{"═"*82}╝')

    if not resultats:
        print('\n  Aucune donnée disponible.\n')
        return

    sep = '─' * 84
    print(f'\n  {"Protocole":<14} {"Marché":<22} '
          f'{"APY Prêteur":>11} {"APY Emprunt.":>13} '
          f'{"Util.":>7} {"TVL":>9} {"Liquidité":>10}')
    print(f'  {sep}')

    for r in sorted(resultats, key=lambda x: x["supply_apy"], reverse=True):
        fu   = " ◄" if r["utilisation"] >= SEUIL_UTIL_ALERTE    else "  "
        fa   = " ★" if r["supply_apy"]  >= SEUIL_APY_OPPORTUNITE else "  "
        ftvl = " ⚠" if r.get("tvl_faible")                       else "  "
        print(f'  {r["protocole"]:<14} {r["marche"]:<22} '
              f'{r["supply_apy"]:>9.2f}%{fa} '
              f'{r["borrow_apy"]:>11.2f}%  '
              f'{r["utilisation"]:>6.1f}%{fu} '
              f'{r["tvl_m"]:>7.1f}M{ftvl} '
              f'{r["liquidite_m"]:>9.1f}M')

    print(f'  {sep}')
    print(f'  ★ APY ≥ {SEUIL_APY_OPPORTUNITE}%  '
          f'◄ Utilisation ≥ {SEUIL_UTIL_ALERTE}%  '
          f'⚠ TVL < {TVL_MIN_M}M (marché peu actif)\n')
    print(f'  Source : smart contracts Ethereum (Alchemy)\n')

    tries = sorted(resultats, key=lambda x: x["supply_apy"], reverse=True)
    if len(tries) >= 2:
        best   = tries[0]
        spread = tries[0]["supply_apy"] - tries[-1]["supply_apy"]
        print(f'  ── Analyse ──────────────────────────────────────────────────')
        print(f'  Meilleur rendement : {best["protocole"]} — {best["marche"]}')
        print(f'    APY prêteur      : {best["supply_apy"]}%')
        print(f'    APY emprunteur   : {best["borrow_apy"]}%')
        print(f'    Liquidité dispo  : {best["liquidite_m"]} M USD')
        print(f'  Spread max / min   : {spread:.2f} pp\n')

    alertes = (
        [f"🔴 {r['protocole']} {r['marche']} : utilisation {r['utilisation']}%"
         for r in resultats if r["utilisation"] >= SEUIL_UTIL_ALERTE] +
        [f"🟡 {r['protocole']} {r['marche']} : APY élevé {r['supply_apy']}%"
         for r in resultats if r["supply_apy"] >= SEUIL_APY_OPPORTUNITE]
    )
    print(f'  ── Alertes ──────────────────────────────────────────────────')
    for a in alertes:
        print(f'  {a}')
    if not alertes:
        print('  ✅ Aucune alerte.')
    print()


# =============================================================================
#  EXPORT EXCEL
# =============================================================================

def maj_excel(resultats, timestamp):
    BLEU  = "1A1A2E"
    GRIS  = "F5F5F5"
    BLANC = "FFFFFF"
    ORG   = "FDF3E8"
    VERT  = "E8F5ED"

    def hdr(cell, bg=BLEU):
        t = Side(style="thin", color="CCCCCC")
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(left=t, right=t, top=t, bottom=t)

    def cel(cell, bg=BLANC, align="left"):
        t = Side(style="thin", color="CCCCCC")
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = Border(left=t, right=t, top=t, bottom=t)

    try:
        wb = load_workbook(FICHIER_EXCEL)
    except FileNotFoundError:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Onglet "Dernière lecture" (réécrit à chaque cycle) ───────────────────
    NOM = "Dernière lecture"
    if NOM in wb.sheetnames:
        del wb[NOM]
    ws = wb.create_sheet(NOM, 0)

    ws["A1"] = f"MONITORING DeFi — On-chain direct — {timestamp}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    ws["A1"].fill      = PatternFill("solid", fgColor=BLEU)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:H1")

    for c, h in enumerate(["Protocole","Marché","APY Prêteur %",
                            "APY Emprunteur %","Utilisation %",
                            "TVL (M$)","Liquidité (M$)","Statut"], 1):
        hdr(ws.cell(row=2, column=c, value=h))

    tries = sorted(resultats, key=lambda x: x["supply_apy"], reverse=True)
    for i, r in enumerate(tries, 3):
        bg = BLANC if i % 2 == 1 else GRIS
        if r.get("tvl_faible"):
            statut, bgs = "TVL < 1M — peu actif", "F0F0F0"
        elif r["utilisation"] >= SEUIL_UTIL_ALERTE:
            statut, bgs = "Utilisation elevee", ORG
        elif r["supply_apy"] >= SEUIL_APY_OPPORTUNITE:
            statut, bgs = "Opportunite", VERT
        else:
            statut, bgs = "Normal", bg
        for c, v in enumerate([r["protocole"], r["marche"], r["supply_apy"],
                                r["borrow_apy"], r["utilisation"],
                                r["tvl_m"], r["liquidite_m"], statut], 1):
            cel(ws.cell(row=i, column=c, value=v),
                bg=bgs if c == 8 else bg,
                align="center" if c > 2 else "left")

    for c, w in enumerate([18,22,14,16,13,10,13,18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Onglet "Historique" (ajout à la suite) ────────────────────────────────
    NOM_H = "Historique"
    if NOM_H not in wb.sheetnames:
        wh = wb.create_sheet(NOM_H)
        for c, h in enumerate(["Timestamp","Protocole","Marché",
                                "APY Prêteur %","APY Emprunteur %",
                                "Utilisation %","TVL (M$)","Liquidité (M$)"], 1):
            hdr(wh.cell(row=1, column=c, value=h))
        for c, w in enumerate([20,14,22,14,16,13,10,13], 1):
            wh.column_dimensions[get_column_letter(c)].width = w
    else:
        wh = wb[NOM_H]

    next_row = wh.max_row + 1
    for r in resultats:
        bg = BLANC if next_row % 2 == 0 else GRIS
        for c, v in enumerate([timestamp, r["protocole"], r["marche"],
                                r["supply_apy"], r["borrow_apy"],
                                r["utilisation"], r["tvl_m"],
                                r["liquidite_m"]], 1):
            cel(wh.cell(row=next_row, column=c, value=v),
                bg=bg, align="center" if c > 3 else "left")
        next_row += 1

    wb.save(FICHIER_EXCEL)
    nb = wh.max_row - 1
    print(f'  📊 Excel → {FICHIER_EXCEL}  '
          f'({nb} observation{"s" if nb > 1 else ""} dans l\'historique)\n')


# =============================================================================
#  SAUVEGARDE CSV
# =============================================================================

def sauvegarder_csv(resultats, timestamp):
    existe = os.path.exists(FICHIER_CSV)
    champs = ['timestamp','protocole','marche','supply_apy',
              'borrow_apy','utilisation','tvl_m','liquidite_m']
    with open(FICHIER_CSV, 'a', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=champs)
        if not existe:
            w.writeheader()
        for r in resultats:
            w.writerow({'timestamp': timestamp,
                        **{k: r.get(k,'') for k in champs[1:]}})


# =============================================================================
#  PROGRAMME PRINCIPAL
# =============================================================================

def connecter():
    load_dotenv()
    url = os.getenv('ALCHEMY_URL')
    if not url:
        print('\n  ❌ ALCHEMY_URL introuvable dans .env\n')
        sys.exit(1)
    w3 = Web3(Web3.HTTPProvider(url))
    if not w3.is_connected():
        print('\n  ❌ Connexion Alchemy échouée.\n')
        sys.exit(1)
    return w3


def run_once(w3, morpho, morpho_params, irm, aave):
    ts   = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    bloc = w3.eth.block_number

    print(f'\n  Lecture des smart contracts...', end=' ', flush=True)
    resultats  = lire_morpho(w3, morpho, morpho_params, irm)
    resultats += lire_aave(w3, aave)
    print(f'OK ({len(resultats)} marchés lus)')

    afficher(resultats, bloc, ts)
    maj_excel(resultats, ts)
    sauvegarder_csv(resultats, ts)


def main():
    w3 = connecter()

    # Un seul contrat Morpho Blue, deux ABIs différents
    morpho = w3.eth.contract(
        address=Web3.to_checksum_address(MORPHO_BLUE_ADDR),
        abi=ABI_MORPHO_MARKET
    )
    morpho_params = w3.eth.contract(
        address=Web3.to_checksum_address(MORPHO_BLUE_ADDR),
        abi=ABI_MORPHO_PARAMS
    )
    irm = w3.eth.contract(
        address=Web3.to_checksum_address(MORPHO_IRM_ADDR),
        abi=ABI_MORPHO_IRM
    )
    aave = w3.eth.contract(
        address=Web3.to_checksum_address(AAVE_PROVIDER_ADDR),
        abi=ABI_AAVE
    )

    if len(sys.argv) > 1 and sys.argv[1] == 'once':
        run_once(w3, morpho, morpho_params, irm, aave)
        return

    print('\n' + '═'*84)
    print('  MONITORING ON-CHAIN DÉMARRÉ')
    print(f'  Rafraîchissement toutes les {INTERVALLE_SECONDES}s — Ctrl+C pour arrêter')
    print(f'  Excel mis à jour automatiquement → {FICHIER_EXCEL}')
    print('═'*84)

    while True:
        try:
            run_once(w3, morpho, morpho_params, irm, aave)
            print(f'  ⏱  Prochain cycle dans {INTERVALLE_SECONDES}s\n')
            time.sleep(INTERVALLE_SECONDES)
        except KeyboardInterrupt:
            print('\n\n  Monitoring arrêté.\n')
            break
        except Exception as e:
            print(f'\n  ⚠️  Erreur : {e}\n')
            time.sleep(INTERVALLE_SECONDES)


if __name__ == '__main__':
    main()
